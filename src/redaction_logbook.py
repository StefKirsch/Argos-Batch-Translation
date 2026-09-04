from __future__ import annotations

from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font


FIELDNAMES = (
    "filename",
    "original version",
    "redacted version",
    "reason for change",
)
DEFAULT_REASON = "Wrong translation in context"
CONTEXT_WORDS = 3
PASSAGE_BREAK_WORDS = 6


def split_words(text: str) -> list[str]:
    """Split text on whitespace while keeping punctuation with its word."""
    return text.split()


def changed_passages(
    original_text: str,
    redacted_text: str,
) -> list[tuple[str, str]]:
    """Return compact word-level snippets for each group of changes.

    A snippet includes up to three words before its first change and ends at the
    last changed word. Six or more unchanged words between changes start a new
    snippet.
    """
    original = split_words(original_text)
    redacted = split_words(redacted_text)
    matcher = SequenceMatcher(None, original, redacted, autojunk=False)
    changes: list[tuple[str, str]] = []
    current_group: list[int] | None = None

    def add_current_group() -> None:
        if current_group is None:
            return

        original_start, original_end, redacted_start, redacted_end = current_group
        original_context_start = max(0, original_start - CONTEXT_WORDS)
        redacted_context_start = max(0, redacted_start - CONTEXT_WORDS)
        changes.append((
            " ".join(original[original_context_start:original_end]),
            " ".join(redacted[redacted_context_start:redacted_end]),
        ))

    for (
        operation,
        original_start,
        original_end,
        redacted_start,
        redacted_end,
    ) in matcher.get_opcodes():
        if operation == "equal":
            if (
                current_group is not None
                and original_end - original_start >= PASSAGE_BREAK_WORDS
            ):
                add_current_group()
                current_group = None
            continue

        if current_group is None:
            current_group = [
                original_start,
                original_end,
                redacted_start,
                redacted_end,
            ]
        else:
            current_group[1] = original_end
            current_group[3] = redacted_end

    add_current_group()

    return changes


def _existing_reasons(logbook_path: Path) -> dict[tuple[str, str, str], str]:
    if not logbook_path.exists():
        return {}

    workbook = load_workbook(logbook_path, read_only=True, data_only=False)
    try:
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None or tuple(header[:len(FIELDNAMES)]) != FIELDNAMES:
            return {}

        reasons: dict[tuple[str, str, str], str] = {}
        for row in rows:
            values = tuple("" if value is None else str(value) for value in row)
            if len(values) < len(FIELDNAMES):
                continue
            reasons[values[:3]] = values[3] or DEFAULT_REASON
        return reasons
    finally:
        workbook.close()


def generate_redaction_logbook(
    translated_dir: Path,
    redacted_dir: Path,
    logbook_path: Path,
) -> tuple[int, list[str]]:
    """Compare translated and redacted documents and write an XLSX logbook.

    Redacted files are matched to translated files by their path relative to the
    respective corpus directory. The return value contains the number of logged
    changes and redacted filenames for which no translated original was found.
    """
    existing_reasons = _existing_reasons(logbook_path)
    rows: list[dict[str, str]] = []
    unmatched_files: list[str] = []

    for redacted_path in sorted(redacted_dir.rglob("*.txt")):
        relative_path = redacted_path.relative_to(redacted_dir)
        filename = relative_path.as_posix()
        original_path = translated_dir / relative_path

        if not original_path.is_file():
            unmatched_files.append(filename)
            continue

        original_text = original_path.read_text(encoding="utf-8")
        redacted_text = redacted_path.read_text(encoding="utf-8")

        for original_passage, redacted_passage in changed_passages(
            original_text,
            redacted_text,
        ):
            key = (filename, original_passage, redacted_passage)
            rows.append({
                "filename": filename,
                "original version": original_passage,
                "redacted version": redacted_passage,
                "reason for change": existing_reasons.get(key, DEFAULT_REASON),
            })

    logbook_path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = logbook_path.with_name(f"{logbook_path.stem}.tmp.xlsx")
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Redaction logbook"
    worksheet.append(FIELDNAMES)
    for row in rows:
        worksheet.append(tuple(row[fieldname] for fieldname in FIELDNAMES))

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.column_dimensions["A"].width = 35
    worksheet.column_dimensions["B"].width = 60
    worksheet.column_dimensions["C"].width = 60
    worksheet.column_dimensions["D"].width = 35
    for row in worksheet.iter_rows():
        for cell in row:
            cell.data_type = "s"
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    workbook.save(temporary_path)
    workbook.close()
    temporary_path.replace(logbook_path)

    return len(rows), unmatched_files
